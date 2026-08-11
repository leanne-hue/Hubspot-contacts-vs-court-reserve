"""
Builds the dashboard dataset from HubSpot contacts + Court Reserve member sets
(read from CSVs), writes one PUBLIC xlsx per location ("Not in Court Reserve")
into docs/downloads/ so it's served directly by GitHub Pages, and renders the
self-contained dashboard HTML.

Locations are inferred from whatever Court Reserve CSVs exist -- not hardcoded.

Court Reserve coverage is computed once per location, against that location's
own "source" of HubSpot contacts and own Court Reserve CSV only
("self_coverage"). This is what the dashboard's "Court Reserve coverage"
table renders -- it does not change based on which location is selected in
the dropdown.

Normally a location's "source" is its Contact Owner (e.g. Vaughan ->
vaughan@pickleplex.ca). But a location that hasn't opened yet has no real
Contact Owner assignments -- config.LOCATION_LIST_SOURCE lets such a location
use a HubSpot list/segment (matched by contact ID) as its source instead, so
its coverage is "list members already in Court Reserve" rather than
"owner's contacts already in Court Reserve".
"""
import os
import json
from collections import defaultdict
from datetime import datetime
from zoneinfo import ZoneInfo

import openpyxl
from openpyxl.styles import Font, PatternFill

from config import derive_owner_email, SITE_DIR, DOWNLOADS_SUBDIR, LOCATION_LIST_SOURCE, NO_CR_LOCATIONS

TZ = ZoneInfo("America/Toronto")


def _safe(name):
    return name.replace(" ", "_").replace("/", "-")


def build_dataset(contacts, cr_members_by_location, cr_last_uploaded="unknown", list_members=None):
    """
    contacts: list of {id,first,last,email,owner_email,owner_name}
    cr_members_by_location: {location: set(emails)}  (locations inferred from CSVs)
    list_members: {location: set(contact_id)} for locations configured in
        config.LOCATION_LIST_SOURCE (e.g. pre-opening mailing-list locations).
    Returns (data_for_dashboard, not_in_cr_rows). Embedded data = aggregate only
    (no names/emails) -- the per-location xlsx files (written separately by
    write_excels) are what carry the actual contact details, and those are
    published as public static files under docs/downloads/.
    """
    list_members = list_members or {}
    locations = sorted(cr_members_by_location.keys())
    owner_map = {loc: derive_owner_email(loc) for loc in locations}

    by_owner = defaultdict(list)
    for c in contacts:
        by_owner[c["owner_email"]].append(c)

    by_id = {c["id"]: c for c in contacts if c.get("id")}

    owner_name_for = {}
    for c in contacts:
        if c["owner_email"] and c["owner_email"] not in owner_name_for:
            owner_name_for[c["owner_email"]] = c.get("owner_name") or c["owner_email"]

    contacts_by_owner = []
    for email, lst in by_owner.items():
        contacts_by_owner.append({
            "owner_name": owner_name_for.get(email, email or "(Unassigned)"),
            "owner_email": email or "",
            "count": len(lst),
        })
    contacts_by_owner.sort(key=lambda r: r["count"], reverse=True)
    total_contacts = len(contacts)

    by_location = {}
    not_in_cr_rows = {}
    self_coverage = []

    for loc in locations:
        cr_emails = cr_members_by_location.get(loc, set())
        mapped_owner = owner_map[loc]

        if loc in LOCATION_LIST_SOURCE:
            # Pre-opening / list-based location: source of truth is HubSpot
            # list membership (matched by contact ID), not Contact Owner.
            member_ids = list_members.get(loc, set())
            source_contacts = [by_id[i] for i in member_ids if i in by_id]
            source = "list"
            source_label = "Pre-Opening Mailing List"
        else:
            source_contacts = by_owner.get(mapped_owner, [])
            source = "owner"
            source_label = owner_name_for.get(mapped_owner, mapped_owner)

        in_cr = sum(1 for c in source_contacts if c["email"] and c["email"] in cr_emails)
        n = len(source_contacts)
        missing = [c for c in source_contacts if not c["email"] or c["email"] not in cr_emails]
        not_in_cr_rows[loc] = missing

        # This location's TRUE coverage: its own source contacts vs its own
        # Court Reserve member list. Nothing here depends on any other
        # location, so it's safe to render as a static table.
        self_coverage.append({
            "location": loc,
            "source": source,
            "source_label": source_label,
            "hubspot_contacts": n,
            "in_cr": in_cr,
            "pct": round(100.0 * in_cr / n, 1) if n else 0.0,
        })

        by_location[loc] = {
            "owner_email": mapped_owner,
            "owner_name": owner_name_for.get(mapped_owner, mapped_owner),
            "source": source,
            "source_label": source_label,
            "hubspot_source_count": n,
            "cr_member_count": len(cr_emails),
            "not_in_cr_count": len(missing),
            "excel_file": f"{DOWNLOADS_SUBDIR}/{_safe(loc)}.xlsx",
        }

    self_coverage.sort(key=lambda r: r["hubspot_contacts"], reverse=True)

    # Locations with no Court Reserve account yet: no coverage to compute,
    # just report the raw HubSpot contact count for that location's owner.
    no_cr_locations = []
    for loc in NO_CR_LOCATIONS:
        mapped_owner = derive_owner_email(loc)
        source_contacts = by_owner.get(mapped_owner, [])
        no_cr_locations.append({
            "location": loc,
            "owner_email": mapped_owner,
            "owner_name": owner_name_for.get(mapped_owner, mapped_owner),
            "hubspot_contacts": len(source_contacts),
        })

    data = {
        "generated_at": datetime.now(TZ).strftime("%A, %B %-d, %Y at %-I:%M %p %Z"),
        "cr_last_uploaded": cr_last_uploaded,
        "locations": locations,
        "owner_map": owner_map,
        "total_contacts": total_contacts,
        "contacts_by_owner": contacts_by_owner,
        "by_location": by_location,
        "self_coverage": self_coverage,
        "no_cr_locations": no_cr_locations,
    }
    return data, not_in_cr_rows


def write_excels(not_in_cr_rows, out_dir=None):
    """Writes one public xlsx per location into docs/downloads/ (or out_dir if given)."""
    if out_dir is None:
        out_dir = os.path.join(SITE_DIR, DOWNLOADS_SUBDIR)
    os.makedirs(out_dir, exist_ok=True)
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1F6F43")
    for loc, rows in not_in_cr_rows.items():
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Not in Court Reserve"
        ws.append(["First Name", "Last Name", "Email", "Contact Owner"])
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
        for c in rows:
            ws.append([c.get("first", ""), c.get("last", ""), c.get("email", ""),
                       c.get("owner_email") or derive_owner_email(loc)])
        for col, width in zip("ABCD", (20, 20, 34, 28)):
            ws.column_dimensions[col].width = width
        ws.freeze_panes = "A2"
        wb.save(os.path.join(out_dir, f"{_safe(loc)}.xlsx"))


def render_html(data) -> str:
    return HTML_TEMPLATE.replace("/*__DATA__*/", json.dumps(data, ensure_ascii=False))


from dashboard_template import HTML_TEMPLATE  # noqa: E402
